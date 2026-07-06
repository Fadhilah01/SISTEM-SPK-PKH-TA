"""
Bulk Import / Export service untuk data calon penerima PKH.

Menyediakan:
  1. Import dari file Excel/CSV → batch insert + prediksi otomatis
  2. Export ke Excel/CSV dengan filter dinamis (rentang waktu, hasil, kolom)
  3. Download template Excel untuk panduan import

Fungsi-fungsi ini dipisahkan dari route handler agar bisa di-test
dan digunakan ulang dari berbagai blueprint.
"""
import io
import os
import re
from datetime import datetime, date, timezone
from collections import OrderedDict

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from sqlalchemy import or_
from models_db import db, CalonPenerima, HasilKeputusan
from core.constants import PENGHASILAN_MAPPING, PEKERJAAN_MAPPING, ASET_MAPPING, KOMPONEN_SOSIAL
from core.scoring import compute_scores, predict_single, create_hasil_keputusan


# ─── Konfigurasi Kolom Template ───

TEMPLATE_COLUMNS = OrderedDict([
    ('nama', 'Nama Kepala Keluarga (wajib)'),
    ('alamat', 'Alamat Lengkap / Detail (wajib)'),
    ('provinsi', 'Provinsi (opsional)'),
    ('kabupaten', 'Kabupaten (opsional)'),
    ('kecamatan', 'Kecamatan (opsional)'),
    ('desa_kelurahan', 'Desa/Kelurahan (opsional)'),
    ('penghasilan', 'Penghasilan per Bulan: ' + ', '.join(PENGHASILAN_MAPPING.keys())),
    ('pekerjaan', 'Pekerjaan: ' + ', '.join(PEKERJAAN_MAPPING.keys())),
    ('kepemilikan_aset', 'Aset: ' + ', '.join(ASET_MAPPING.keys())),
    ('ibu_hamil', 'Ibu Hamil? (YA/TIDAK atau 1/0)'),
    ('anak_usia_dini', 'Anak Usia Dini? (YA/TIDAK atau 1/0)'),
    ('anak_sekolah', 'Anak Sekolah? (YA/TIDAK atau 1/0)'),
    ('disabilitas', 'Disabilitas? (YA/TIDAK atau 1/0)'),
    ('lansia', 'Lansia? (YA/TIDAK atau 1/0)'),
])

REQUIRED_COLUMNS = {'nama', 'alamat', 'penghasilan', 'pekerjaan', 'kepemilikan_aset'}

# Mapping teks YA/TIDAK ke boolean
_BOOL_MAP = {
    'ya': True, 'y': True, '1': True, 'yes': True, 'true': True, 'ada': True, '✔': True,
    'tidak': False, 't': False, '0': False, 'no': False, 'false': False, 'none': False, '': False,
}


def parse_bool(val):
    """Konversi berbagai format input ke boolean. Case-insensitive."""
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return val != 0
    if isinstance(val, str):
        return _BOOL_MAP.get(val.strip().lower(), False)
    return False


def validate_row(row_dict, row_num):
    """
    Validasi satu baris data dari file.

    Returns:
        (is_valid: bool, errors: list[str])
    """
    errors = []

    # Wajib: nama
    nama = str(row_dict.get('nama', '') or '').strip()
    if not nama:
        errors.append(f"Baris {row_num}: Nama tidak boleh kosong")

    # Wajib: alamat
    alamat = str(row_dict.get('alamat', '') or '').strip()
    if not alamat:
        errors.append(f"Baris {row_num}: Alamat tidak boleh kosong")

    # Validasi penghasilan
    penghasilan = str(row_dict.get('penghasilan', '') or '').strip()
    if penghasilan not in PENGHASILAN_MAPPING:
        errors.append(
            f"Baris {row_num}: Penghasilan '{penghasilan}' tidak valid. "
            f"Pilih: {', '.join(PENGHASILAN_MAPPING.keys())}"
        )

    # Validasi pekerjaan
    pekerjaan = str(row_dict.get('pekerjaan', '') or '').strip()
    if pekerjaan not in PEKERJAAN_MAPPING:
        errors.append(
            f"Baris {row_num}: Pekerjaan '{pekerjaan}' tidak valid. "
            f"Pilih: {', '.join(PEKERJAAN_MAPPING.keys())}"
        )

    # Validasi aset
    aset = str(row_dict.get('kepemilikan_aset', '') or '').strip()
    if aset not in ASET_MAPPING:
        errors.append(
            f"Baris {row_num}: Aset '{aset}' tidak valid. "
            f"Pilih: {', '.join(ASET_MAPPING.keys())}"
        )

    return len(errors) == 0, errors


def row_to_calon_data(row_dict, predictor):
    """
    Konversi satu baris pandas Series/dict ke data calon.
    Langsung hitung skor + prediksi.

    Returns:
        dict dengan keys: calon_dict, skor, result (None jika predictor=None)
        atau None jika gagal
    """
    nama = str(row_dict.get('nama', '') or '').strip()
    alamat = str(row_dict.get('alamat', '') or '').strip()
    
    # Pengambilan opsional kolom wilayah
    provinsi = str(row_dict.get('provinsi', '') or '').strip() or None
    kabupaten = str(row_dict.get('kabupaten', '') or '').strip() or None
    kecamatan = str(row_dict.get('kecamatan', '') or '').strip() or None
    desa_kelurahan = str(row_dict.get('desa_kelurahan', '') or '').strip() or None

    penghasilan = str(row_dict.get('penghasilan', '') or '').strip()
    pekerjaan = str(row_dict.get('pekerjaan', '') or '').strip()
    aset = str(row_dict.get('kepemilikan_aset', '') or '').strip()

    # boolean fields
    ibu_hamil = parse_bool(row_dict.get('ibu_hamil'))
    anak_usia_dini = parse_bool(row_dict.get('anak_usia_dini'))
    anak_sekolah = parse_bool(row_dict.get('anak_sekolah'))
    disabilitas = parse_bool(row_dict.get('disabilitas'))
    lansia = parse_bool(row_dict.get('lansia'))

    skor = compute_scores(penghasilan, pekerjaan, aset,
                          ibu_hamil=ibu_hamil,
                          anak_usia_dini=anak_usia_dini,
                          anak_sekolah=anak_sekolah,
                          disabilitas=disabilitas,
                          lansia=lansia)

    calon_dict = {
        'nama': nama,
        'alamat': alamat,
        'provinsi': provinsi,
        'kabupaten': kabupaten,
        'kecamatan': kecamatan,
        'desa_kelurahan': desa_kelurahan,
        'penghasilan': penghasilan,
        'pekerjaan': pekerjaan,
        'kepemilikan_aset': aset,
        'ibu_hamil': ibu_hamil,
        'anak_usia_dini': anak_usia_dini,
        'anak_sekolah': anak_sekolah,
        'disabilitas': disabilitas,
        'lansia': lansia,
    }
    calon_dict.update(skor)

    result = None
    if predictor:
        # Buat objek dummy untuk predict_single
        class _DummyCalon:
            pass
        dummy = _DummyCalon()
        for k, v in skor.items():
            setattr(dummy, k, v)
        dummy.ibu_hamil = ibu_hamil
        dummy.anak_usia_dini = anak_usia_dini
        dummy.anak_sekolah = anak_sekolah
        dummy.disabilitas = disabilitas
        dummy.lansia = lansia

        result = predict_single(predictor, dummy)

    return {
        'calon_dict': calon_dict,
        'skor': skor,
        'result': result,
    }


def import_from_file(file_storage, predictor):
    """
    Baca file Excel/CSV, validasi baris, simpak ke DB.

    Parameters
    ----------
    file_storage : werkzeug.FileStorage — file upload dari form
    predictor : SVMPredictor | None

    Returns
    -------
    dict : {
        'success': int,       # jumlah berhasil import
        'failed': int,        # jumlah gagal validasi
        'total': int,         # total baris (excl header)
        'errors': list[str],  # daftar pesan error detail
        'imported_ids': list[int]  # ID calon yang berhasil di-import
    }
    """
    filename = file_storage.filename.lower()
    if filename.endswith('.csv'):
        df = pd.read_csv(file_storage)
    else:
        df = pd.read_excel(file_storage, engine='openpyxl')

    # Normalize column names
    col_map = {}
    for col in df.columns:
        col_clean = col.strip().lower().replace(' ', '_').replace('-', '_')
        col_map[col] = col_clean
    df = df.rename(columns=col_map)

    # Track columns found vs required (Hanya yang benar-benar wajib)
    required = REQUIRED_COLUMNS
    found = set(df.columns)
    missing = required - found

    if missing:
        return {
            'success': 0,
            'failed': 0,
            'total': len(df),
            'errors': [f"Kolom wajib tidak ditemukan: {', '.join(sorted(missing))}. "
                       f"Gunakan template yang disediakan."],
            'imported_ids': [],
        }

    hasil = {
        'success': 0,
        'failed': 0,
        'total': len(df),
        'errors': [],
        'imported_ids': [],
    }

    for idx, row in df.iterrows():
        row_num = idx + 2  # +1 untuk header, +1 untuk 0-index
        row_dict = row.to_dict()

        is_valid, errs = validate_row(row_dict, row_num)
        if not is_valid:
            hasil['failed'] += 1
            hasil['errors'].extend(errs)
            continue

        try:
            data = row_to_calon_data(row_dict, predictor)

            calon = CalonPenerima(**data['calon_dict'])
            db.session.add(calon)
            db.session.flush()  # dapatkan ID

            if data['result']:
                kep_data = create_hasil_keputusan(
                    calon.id, data['result'],
                    oleh='Import Bulk'
                )
                db.session.add(HasilKeputusan(**kep_data))

            hasil['success'] += 1
            hasil['imported_ids'].append(calon.id)

        except Exception as e:
            db.session.rollback()
            hasil['failed'] += 1
            hasil['errors'].append(f"Baris {row_num}: {str(e)}")
            # Re-begin transaction for next row
            continue

    db.session.commit()
    return hasil


# ─── Export Service ───

def export_data(format_type='excel', filters=None):
    """
    Export data calon + hasil keputusan ke file Excel atau CSV.

    Parameters
    ----------
    format_type : str — 'excel' atau 'csv'
    filters : dict — filter opsional:
        - date_from: str (YYYY-MM-DD)
        - date_to: str (YYYY-MM-DD)
        - hasil: 'layak' | 'tidak_layak' | '' (semua)
        - columns: list[str] — subset kolom (default: semua)
        - q: str — filter nama/alamat

    Returns
    -------
    tuple (io.BytesIO|str, filename, mime_type)
        - file_obj / string data
        - suggested filename
        - content type
    """
    if filters is None:
        filters = {}

    # ── Query data ──
    query = (
        db.session.query(CalonPenerima, HasilKeputusan)
        .outerjoin(HasilKeputusan, CalonPenerima.id == HasilKeputusan.id_calon)
    )

    # Filter nama/alamat
    if filters.get('q'):
        q = filters['q']
        query = query.filter(
            or_(
                CalonPenerima.nama.like(f'%{q}%'),
                CalonPenerima.alamat.like(f'%{q}%'),
            )
        )

    # Filter tanggal
    date_from = filters.get('date_from')
    date_to = filters.get('date_to')

    if date_from:
        try:
            dt_from = datetime.strptime(date_from, '%Y-%m-%d').replace(tzinfo=timezone.utc)
            query = query.filter(CalonPenerima.created_at >= dt_from)
        except (ValueError, TypeError):
            pass

    if date_to:
        try:
            dt_to = datetime.strptime(date_to, '%Y-%m-%d').replace(
                hour=23, minute=59, second=59, tzinfo=timezone.utc
            )
            query = query.filter(CalonPenerima.created_at <= dt_to)
        except (ValueError, TypeError):
            pass

    # Filter hasil
    hasil_filter = filters.get('hasil', '')
    if hasil_filter == 'layak':
        query = query.filter(HasilKeputusan.hasil_prediksi == True)
    elif hasil_filter == 'tidak_layak':
        query = query.filter(HasilKeputusan.hasil_prediksi == False)

    query = query.order_by(CalonPenerima.id.asc())
    rows = query.all()

    # ── Build DataFrame ──
    all_columns = [
        ('ID', lambda c, h: c.id),
        ('Nama', lambda c, h: c.nama),
        ('Alamat', lambda c, h: c.alamat),
        ('Provinsi', lambda c, h: c.provinsi or ''),
        ('Kabupaten', lambda c, h: c.kabupaten or ''),
        ('Kecamatan', lambda c, h: c.kecamatan or ''),
        ('Desa', lambda c, h: c.desa_kelurahan or ''),
        ('Penghasilan', lambda c, h: c.penghasilan),
        ('Pekerjaan', lambda c, h: c.pekerjaan),
        ('Kepemilikan Aset', lambda c, h: c.kepemilikan_aset),
        ('Ibu Hamil', lambda c, h: 'Ya' if c.ibu_hamil else 'Tidak'),
        ('Anak Usia Dini', lambda c, h: 'Ya' if c.anak_usia_dini else 'Tidak'),
        ('Anak Sekolah', lambda c, h: 'Ya' if c.anak_sekolah else 'Tidak'),
        ('Disabilitas', lambda c, h: 'Ya' if c.disabilitas else 'Tidak'),
        ('Lansia', lambda c, h: 'Ya' if c.lansia else 'Tidak'),
        ('Skor Penghasilan', lambda c, h: c.skor_penghasilan),
        ('Skor Pekerjaan', lambda c, h: c.skor_pekerjaan),
        ('Skor Aset', lambda c, h: c.skor_kepemilikan_aset),
        ('Skor Ibu Hamil', lambda c, h: c.skor_ibu_hamil),
        ('Skor Anak Usia Dini', lambda c, h: c.skor_anak_usia_dini),
        ('Skor Anak Sekolah', lambda c, h: c.skor_anak_sekolah),
        ('Skor Disabilitas', lambda c, h: c.skor_disabilitas),
        ('Skor Lansia', lambda c, h: c.skor_lansia),
        ('Keputusan', lambda c, h: h.label_prediksi if h else 'Belum Diprediksi'),
        ('Probabilitas (%)', lambda c, h: round(h.probabilitas * 100, 2) if h and h.probabilitas is not None else ''),
        ('Tanggal Input', lambda c, h: c.created_at.strftime('%d/%m/%Y %H:%M') if c.created_at else ''),
        ('Tanggal Prediksi', lambda c, h: h.tanggal_prediksi.strftime('%d/%m/%Y %H:%M') if h and h.tanggal_prediksi else ''),
        ('Oleh', lambda c, h: h.oleh if h else ''),
    ]

    # Filter selected columns
    selected_cols = filters.get('columns')
    if selected_cols:
        # Map column keys to display names
        col_key_map = {
            'id': 'ID', 'nama': 'Nama', 'alamat': 'Alamat',
            'provinsi': 'Provinsi', 'kabupaten': 'Kabupaten',
            'kecamatan': 'Kecamatan', 'desa_kelurahan': 'Desa',
            'penghasilan': 'Penghasilan', 'pekerjaan': 'Pekerjaan',
            'kepemilikan_aset': 'Kepemilikan Aset',
            'ibu_hamil': 'Ibu Hamil', 'anak_usia_dini': 'Anak Usia Dini',
            'anak_sekolah': 'Anak Sekolah', 'disabilitas': 'Disabilitas',
            'lansia': 'Lansia',
            'skor_penghasilan': 'Skor Penghasilan', 'skor_pekerjaan': 'Skor Pekerjaan',
            'skor_kepemilikan_aset': 'Skor Aset',
            'skor_ibu_hamil': 'Skor Ibu Hamil', 'skor_anak_usia_dini': 'Skor Anak Usia Dini',
            'skor_anak_sekolah': 'Skor Anak Sekolah',
            'skor_disabilitas': 'Skor Disabilitas', 'skor_lansia': 'Skor Lansia',
            'hasil_prediksi': 'Keputusan', 'probabilitas': 'Probabilitas (%)',
            'tanggal_input': 'Tanggal Input', 'tanggal_prediksi': 'Tanggal Prediksi',
            'oleh': 'Oleh',
        }
        display_names = [col_key_map.get(c, c) for c in selected_cols]
        all_columns = [(name, fn) for name, fn in all_columns if name in display_names]

    data = []
    for c, h in rows:
        row_data = {}
        for name, fn in all_columns:
            try:
                row_data[name] = fn(c, h)
            except Exception:
                row_data[name] = ''
        data.append(row_data)

    df_export = pd.DataFrame(data)

    # Generate filename with timestamp
    now_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    filter_tag = ''
    if hasil_filter:
        filter_tag = f'_{hasil_filter}'

    if format_type == 'csv':
        buf = io.BytesIO()
        df_export.to_csv(buf, index=False, encoding='utf-8-sig')
        buf.seek(0)
        filename = f'data_pkh{filter_tag}_{now_str}.csv'
        mime_type = 'text/csv; charset=utf-8-sig'
        return buf, filename, mime_type
    else:
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as writer:
            df_export.to_excel(writer, index=False, sheet_name='Data PKH')
        buf.seek(0)
        filename = f'data_pkh{filter_tag}_{now_str}.xlsx'
        mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return buf, filename, mime_type


def get_column_options():
    """Daftar kolom untuk form export filter (checkbox dinamis)."""
    return [
        {'key': 'id', 'label': 'ID'},
        {'key': 'nama', 'label': 'Nama'},
        {'key': 'alamat', 'label': 'Alamat'},
        {'key': 'provinsi', 'label': 'Provinsi'},
        {'key': 'kabupaten', 'label': 'Kabupaten'},
        {'key': 'kecamatan', 'label': 'Kecamatan'},
        {'key': 'desa_kelurahan', 'label': 'Desa/Kelurahan'},
        {'key': 'penghasilan', 'label': 'Penghasilan'},
        {'key': 'pekerjaan', 'label': 'Pekerjaan'},
        {'key': 'kepemilikan_aset', 'label': 'Kepemilikan Aset'},
        {'key': 'ibu_hamil', 'label': 'Ibu Hamil'},
        {'key': 'anak_usia_dini', 'label': 'Anak Usia Dini'},
        {'key': 'anak_sekolah', 'label': 'Anak Sekolah'},
        {'key': 'disabilitas', 'label': 'Disabilitas'},
        {'key': 'lansia', 'label': 'Lansia'},
        {'key': 'skor_penghasilan', 'label': 'Skor Penghasilan'},
        {'key': 'skor_pekerjaan', 'label': 'Skor Pekerjaan'},
        {'key': 'skor_kepemilikan_aset', 'label': 'Skor Aset'},
        {'key': 'skor_ibu_hamil', 'label': 'Skor Ibu Hamil'},
        {'key': 'skor_anak_usia_dini', 'label': 'Skor Anak Usia Dini'},
        {'key': 'skor_anak_sekolah', 'label': 'Skor Anak Sekolah'},
        {'key': 'skor_disabilitas', 'label': 'Skor Disabilitas'},
        {'key': 'skor_lansia', 'label': 'Skor Lansia'},
        {'key': 'hasil_prediksi', 'label': 'Keputusan (Layak/Tidak)'},
        {'key': 'probabilitas', 'label': 'Probabilitas (%)'},
        {'key': 'tanggal_input', 'label': 'Tanggal Input'},
        {'key': 'tanggal_prediksi', 'label': 'Tanggal Prediksi'},
        {'key': 'oleh', 'label': 'Oleh (Importir/Petugas)'},
    ]


# ─── Template Generator ───

def generate_template():
    """
    Generate template Excel (.xlsx) untuk panduan import bulk.

    Returns:
        io.BytesIO — file template siap di-download
    """
    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Template Import PKH'

    # ── Header row ──
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for col_idx, (col_key, col_desc) in enumerate(TEMPLATE_COLUMNS.items(), 1):
        cell = ws.cell(row=1, column=col_idx, value=col_key)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    # ── Row 2: contoh data ──
    sample = [
        'Ahmad Syahputra',      # nama
        'Jalan Sis Aljufri No. 12', # alamat
        'SULAWESI TENGAH',      # provinsi
        'KOTA PALU',            # kabupaten
        'MANTIKULORE',          # kecamatan
        'TONDO',                # desa_kelurahan
        'Desil 1 (< Rp.500.000)',  # penghasilan
        'Tidak Bekerja',        # pekerjaan
        'Tidak Memiliki Aset',  # aset
        'Ya',                   # ibu_hamil
        'Ya',                   # anak_usia_dini
        'Tidak',                # anak_sekolah
        'Tidak',                # disabilitas
        'Tidak',                # lansia
    ]
    example_font = Font(name='Calibri', color='666666', italic=True, size=10)
    for col_idx, val in enumerate(sample, 1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.font = example_font
        cell.alignment = Alignment(vertical='center')
        cell.border = thin_border

    # ── Row 3: Panduan ──
    ws.append([])
    guide_title = ws.cell(row=4, column=1, value='📋 PANDUAN PENGISIAN:')
    guide_title.font = Font(name='Calibri', bold=True, size=11, color='0D6EFD')

    guide_data = [
        '1. Baris pertama (header) HARUS menggunakan nama kolom seperti di atas — jangan diubah.',
        '2. Baris kedua adalah CONTOH — boleh dihapus atau ditimpa dengan data sebenarnya.',
        '3. Kolom penghasilan, pekerjaan, dan kepemilikan_aset harus diisi dengan kategori yang sesuai (lihat daftar di bawah).',
        '4. Kolom boolean (ibu_hamil, anak_usia_dini, dll): isi YA/TIDAK, 1/0, atau Ada/Tidak.',
        '5. Semua data akan diproses dan diprediksi otomatis oleh sistem SVM.',
    ]
    guide_font = Font(name='Calibri', size=10, color='333333')
    for i, line in enumerate(guide_data, 5):
        ws.cell(row=i, column=1, value=line).font = guide_font

    # ── Daftar kategori ──
    cat_row = 11
    ws.cell(row=cat_row, column=1, value='DAFTAR KATEGORI YANG VALID:').font = Font(name='Calibri', bold=True, size=11, color='0D6EFD')

    cat_row += 1
    ws.cell(row=cat_row, column=1, value='Penghasilan:').font = Font(bold=True, size=10)
    for i, k in enumerate(PENGHASILAN_MAPPING.keys(), cat_row + 1):
        ws.cell(row=i, column=1, value=f'  • {k}').font = Font(size=10)

    cat_row += len(PENGHASILAN_MAPPING) + 1
    ws.cell(row=cat_row, column=1, value='Pekerjaan:').font = Font(bold=True, size=10)
    for i, k in enumerate(PEKERJAAN_MAPPING.keys(), cat_row + 1):
        ws.cell(row=i, column=1, value=f'  • {k}').font = Font(size=10)

    cat_row += len(PEKERJAAN_MAPPING) + 1
    ws.cell(row=cat_row, column=1, value='Kepemilikan Aset:').font = Font(bold=True, size=10)
    for i, k in enumerate(ASET_MAPPING.keys(), cat_row + 1):
        ws.cell(row=i, column=1, value=f'  • {k}').font = Font(size=10)

    # Freeze header
    ws.freeze_panes = 'A2'

    # Auto-filter
    ws.auto_filter.ref = f'A1:{get_column_letter(len(TEMPLATE_COLUMNS))}1'

    # ── Sheet 2: Petunjuk Lengkap ──
    ws2 = wb.create_sheet('Petunjuk')
    instructions = [
        'PETUNJUK LENGKAP IMPORT DATA CALON PKH',
        '',
        '1. PERSIAPAN DATA',
        '   - Gunakan template ini sebagai dasar pengisian data.',
        '   - Isi baris pertama (header) JANGAN dihapus atau diubah.',
        '   - Hapus baris contoh (baris 2) dan ganti dengan data real.',
        '',
        '2. ATURAN PENGISIAN',
        '   - Nama: Nama kepala keluarga (max 100 karakter)',
        '   - Alamat: Nama desa/kelurahan (max 255 karakter)',
        '   - Penghasilan: Pilih salah satu dari 5 kategori Desil',
        '   - Pekerjaan: Pilih salah satu dari 5 kategori',
        '   - Kepemilikan Aset: Pilih salah satu dari 5 kategori',
        '   - Komponen Sosial (Ibu Hamil, dll):',
        '     YA / Tidak / 1 / 0 / Ada / Tidak Ada',
        '',
        '3. PROSES IMPORT',
        '   - File akan divalidasi baris per baris.',
        '   - Baris yang valid akan langsung disimpan + diprediksi SVM.',
        '   - Baris yang gagal validasi akan dilewati + dicatat errornya.',
        '   - Hasil akhir: laporan jumlah sukses & gagal.',
    ]
    for i, line in enumerate(instructions, 1):
        ws2.cell(row=i, column=1, value=line).font = Font(name='Calibri', size=11)

    ws2.column_dimensions['A'].width = 80

    wb.save(buf)
    buf.seek(0)
    return buf
